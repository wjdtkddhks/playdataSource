from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws = wb.create_sheet('diary', 0)

data = [('홍길동', 80, 70, 90), ('이기자', 90, 60, 80), ('김기자', 80, 80, 70)]

r = 1 #첫번째 행번호는 1부터 시작
c = 1 #첫번째 열번호는 1부터 시작

for irum, kor, eng, math in data:
    ws.cell(row=r, column=c).value = irum
    ws.cell(row=r, column=c+1).value = kor
    ws.cell(row=r, column=c+2).value = eng
    ws.cell(row=r, column=c+3).value = math
    r += 1

"""
for irum, kor, eng, math in data:
    ws['A'+str(r)].value = irum
    ws['B'+str(r)].value = kor
    ws['C'+str(r)].value = eng
    ws['D'+str(r)].value = math
    r += 1

columnChar = 'A'
for irum, kor, eng, math in data:
    ws[columnChar+str(r)].value = irum
    # ws[columnChar+str(r)].value = irum.encode(encoding='utf-8', errors='ignore')
    ws[chr(ord(columnChar)+1)+str(r)].value = kor # chr(): 유니코드포인트가 정수인 수를 문자로, ord(): 그반대
    ws[chr(ord(columnChar)+2)+str(r)].value = eng
    ws[chr(ord(columnChar)+3)+str(r)].value = math
    r += 1
    
columnChar = 'A'
for val in data:
    for i in range(0,4):
        ws[chr(ord(columnChar)+i)+str(r).value = val[i]
        
    r += 1

"""

ws.cell(row=r, column=1).value = '합계'
ws.cell(row=r, column=2).value = '=sum(B1:B3)'
ws.cell(row=r, column=3).value = '=sum(C1:C3)'
ws.cell(row=r, column=4).value = '=sum(D1:D3)'

wb.save('openpyxl2.xlsx')
wb.close()

