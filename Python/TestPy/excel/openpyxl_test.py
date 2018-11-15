# https://openpyxl.readthedocs.io/en/stable/ 사이트 참조
from openpyxl import Workbook, load_workbook

wb = Workbook() #새문서를 만듬
ws = wb.active #활성화되어 있는 시트 선택
# ws = wb.create_sheet(0) 0이면 첫번째 위치에 삽입, 생략시 마지막 위치에 삽입

ws.title = 'sample'
ws['B2'] = 42
ws.append([1,2,3]) # 가장 마지막에 추가된 다음 행부터 추가
ws.append([1,2,3])
ws.append([1,2,3])
ws.append([4,4,4])
wb.save('openpyxl1.xlsx')
wb.close()

wb = load_workbook(filename='openpyxl1.xlsx')
ws = wb.active
ws['A1'] = 42
ws.cell(row=1, column=3).value = 333
ws.append(['aaa', 'bbb', 'ccc']) #새로운 행에 추가

print("A1:", ws['A1'].value)
print("A2:", ws['A2'].value) #저장된 값이 없으면 None으로 출력

ws2 = wb['sample'] #샘플시트 선택
print(ws2['A3'].value, ws2['B3'].value, ws2['C3'].value)
print(ws2['A4'].value, ws2['B4'].value, ws2['C4'].value)
print(ws2['A5'].value, ws2['B5'].value, ws2['C5'].value)

wb.save('openpyxl1.xlsx')

